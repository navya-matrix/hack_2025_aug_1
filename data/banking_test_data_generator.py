import random
import string
import pandas as pd
from datetime import datetime, timedelta, date
from faker import Faker
import uuid
import openpyxl
from openpyxl.styles import Font, PatternFill
import numpy as np

# Initialize Faker
fake = Faker()

# Set seed for reproducible results
random.seed(42)
fake.seed_instance(42)

class BankingDataGenerator:
    def __init__(self):
        self.branches = []
        self.customers = []
        self.employees = []
        self.accounts = []
        self.transactions = []
        
    def generate_id(self, prefix="", length=8):
        """Generate a unique ID with optional prefix"""
        return prefix + ''.join(random.choices(string.ascii_uppercase + string.digits, k=length))
    
    def generate_branches(self, count=50):
        """Generate branch data"""
        print(f"Generating {count} branches...")
        
        branch_names = [
            "Main Branch", "Downtown Branch", "North Branch", "South Branch", 
            "East Branch", "West Branch", "Airport Branch", "Mall Branch",
            "University Branch", "Corporate Branch", "Suburban Branch", 
            "Financial District Branch", "Residential Branch", "Commercial Branch"
        ]
        
        cities = ["New York", "Los Angeles", "Chicago", "Houston", "Phoenix", 
                 "Philadelphia", "San Antonio", "San Diego", "Dallas", "San Jose",
                 "Austin", "Jacksonville", "Fort Worth", "Columbus", "Indianapolis"]
        
        states = ["NY", "CA", "IL", "TX", "AZ", "PA", "FL", "OH", "IN"]
        
        for i in range(count):
            branch = {
                'id': self.generate_id('BR', 10),
                'name': f"{random.choice(branch_names)} {i+1}" if i >= len(branch_names) else f"{branch_names[i % len(branch_names)]} {(i//len(branch_names))+1}",
                'address': fake.street_address(),
                'city': random.choice(cities),
                'state': random.choice(states),
                'zip_code': fake.zipcode(),
                'manager_id': None,  # Will be set after employees are created
                'created_at': fake.date_time_between(start_date='-5y', end_date='now'),
                'updated_at': fake.date_time_between(start_date='-1y', end_date='now')
            }
            self.branches.append(branch)
    
    def generate_employees(self, count=500):
        """Generate employee data"""
        print(f"Generating {count} employees...")
        
        positions = [
            "Manager", "Assistant Manager", "Teller", "Loan Officer", 
            "Customer Service Representative", "Financial Advisor", 
            "Operations Specialist", "Security Officer", "IT Support",
            "Branch Manager", "Senior Teller", "Credit Analyst"
        ]
        
        for i in range(count):
            branch_id = random.choice(self.branches)['id']
            hire_date = fake.date_between(start_date='-10y', end_date='now')
            
            employee = {
                'id': self.generate_id('EMP', 10),
                'branch_id': branch_id,
                'name': fake.name(),
                'email': fake.unique.email(),
                'phone': fake.phone_number(),
                'position': random.choice(positions),
                'hire_date': hire_date,
                'salary': round(random.uniform(30000, 150000), 2),
                'created_at': fake.date_time_between(start_date=hire_date, end_date='now'),
                'updated_at': fake.date_time_between(start_date=hire_date, end_date='now')
            }
            self.employees.append(employee)
        
        # Assign managers to branches
        managers = [emp for emp in self.employees if emp['position'] in ['Manager', 'Branch Manager']]
        for branch in self.branches:
            available_managers = [m for m in managers if m['branch_id'] == branch['id']]
            if available_managers:
                branch['manager_id'] = random.choice(available_managers)['id']
    
    def generate_customers(self, count=2000):
        """Generate customer data with edge cases"""
        print(f"Generating {count} customers...")
        
        genders = ['male', 'female', 'other', 'prefer_not_to_say']
        
        for i in range(count):
            # Create some edge cases
            birth_date = fake.date_between(start_date='-90y', end_date='-18y')
            
            # Some customers with very old or very recent registration
            if i < 50:  # Very old customers
                created_date = fake.date_time_between(start_date='-20y', end_date='-15y')
            elif i < 100:  # Very recent customers
                created_date = fake.date_time_between(start_date='-30d', end_date='now')
            else:
                created_date = fake.date_time_between(start_date='-10y', end_date='now')
            
            customer = {
                'id': self.generate_id('CUST', 10),
                'email': fake.unique.email(),
                'phone': fake.phone_number() if random.random() > 0.05 else None,  # 5% without phone
                'address': fake.address() if random.random() > 0.02 else None,  # 2% without address
                'first_name': fake.first_name(),
                'last_name': fake.last_name(),
                'date_of_birth': birth_date,
                'gender': random.choice(genders),
                'national_id': fake.unique.ssn(),
                'created_at': created_date,
                'updated_at': fake.date_time_between(start_date=created_date, end_date='now'),
                'branch_id': random.choice(self.branches)['id']
            }
            self.customers.append(customer)
    
    def generate_accounts(self, count=3000):
        """Generate account data with various types and statuses"""
        print(f"Generating {count} accounts...")
        
        account_types = ['checking', 'savings', 'credit', 'loan']
        statuses = ['active', 'inactive', 'closed', 'suspended']
        
        for i in range(count):
            customer = random.choice(self.customers)
            account_type = random.choice(account_types)
            status = random.choice(statuses) if i < 200 else 'active'  # Most accounts active
            
            # Different balance ranges based on account type
            if account_type == 'checking':
                balance = round(random.uniform(-1000, 50000), 2)
            elif account_type == 'savings':
                balance = round(random.uniform(0, 100000), 2)
            elif account_type == 'credit':
                balance = round(random.uniform(-10000, 0), 2)  # Negative for credit
            else:  # loan
                balance = round(random.uniform(-50000, 0), 2)  # Negative for loan
            
            # Some edge cases for testing
            if i < 10:  # Very high balance accounts
                balance = round(random.uniform(1000000, 5000000), 2)
            elif i < 20:  # Zero balance accounts
                balance = 0.00
            elif i < 30:  # Very low balance accounts
                balance = round(random.uniform(0.01, 10.00), 2)
            
            opened_date = fake.date_between(start_date=customer['created_at'].date(), end_date='now')
            
            account = {
                'id': self.generate_id('ACC', 10),
                'customer_id': customer['id'],
                'account_number': fake.unique.bban(),
                'type': account_type,
                'balance': balance,
                'opened_at': opened_date,
                'interest_rate': round(random.uniform(0.01, 5.00), 4) if account_type in ['savings', 'loan'] else None,
                'status': status,
                'branch_id': customer['branch_id'],
                'created_at': fake.date_time_between(start_date=opened_date, end_date='now'),
                'updated_at': fake.date_time_between(start_date=opened_date, end_date='now')
            }
            self.accounts.append(account)
    
    def generate_transactions(self, count=15000):
        """Generate transaction data with various patterns"""
        print(f"Generating {count} transactions...")
        
        transaction_types = ['deposit', 'withdrawal', 'transfer', 'fee', 'interest']
        statuses = ['pending', 'completed', 'failed', 'cancelled']
        
        for i in range(count):
            account = random.choice(self.accounts)
            trans_type = random.choice(transaction_types)
            status = random.choice(statuses) if i < 500 else 'completed'  # Most completed
            
            # Amount based on transaction type
            if trans_type == 'deposit':
                amount = round(random.uniform(10, 10000), 2)
            elif trans_type == 'withdrawal':
                amount = round(random.uniform(-5000, -1), 2)
            elif trans_type == 'transfer':
                amount = round(random.uniform(-2000, 2000), 2)
            elif trans_type == 'fee':
                amount = round(random.uniform(-100, -1), 2)
            else:  # interest
                amount = round(random.uniform(0.01, 500), 2)
            
            # Some edge cases
            if i < 20:  # Very large transactions
                amount = round(random.uniform(100000, 1000000), 2) if trans_type == 'deposit' else round(random.uniform(-1000000, -100000), 2)
            elif i < 40:  # Very small transactions
                amount = round(random.uniform(0.01, 1.00), 2) if trans_type == 'deposit' else round(random.uniform(-1.00, -0.01), 2)
            
            # Transaction date should be after account opening
            start_date = max(account['opened_at'], (datetime.now() - timedelta(days=365)).date())
            trans_date = fake.date_time_between(start_date=start_date, end_date='now')
            
            transaction = {
                'id': self.generate_id('TXN', 10),
                'account_id': account['id'],
                'transaction_date': trans_date,
                'amount': amount,
                'type': trans_type,
                'description': fake.text(max_nb_chars=100),
                'status': status,
                'created_at': trans_date,
                'updated_at': fake.date_time_between(start_date=trans_date, end_date='now'),
                'employee_id': random.choice(self.employees)['id'] if random.random() > 0.3 else None
            }
            self.transactions.append(transaction)
    
    def generate_all_data(self):
        """Generate all test data"""
        print("Starting data generation...")
        self.generate_branches(50)
        self.generate_employees(500)
        self.generate_customers(2000)
        self.generate_accounts(3000)
        self.generate_transactions(15000)
        print("Data generation completed!")
        
        # Print summary
        print(f"\nData Summary:")
        print(f"Branches: {len(self.branches)}")
        print(f"Employees: {len(self.employees)}")
        print(f"Customers: {len(self.customers)}")
        print(f"Accounts: {len(self.accounts)}")
        print(f"Transactions: {len(self.transactions)}")
        print(f"Total records: {len(self.branches) + len(self.employees) + len(self.customers) + len(self.accounts) + len(self.transactions)}")
    
    def generate_sql_inserts(self, filename="banking_test_data_new.sql"):
        """Generate SQL INSERT statements"""
        print(f"Generating SQL INSERT statements to {filename}...")
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write("-- Banking System Test Data\n")
            f.write("-- Generated test data with relational integrity\n\n")
            
            # Branches
            f.write("-- Insert Branches\n")
            for branch in self.branches:
                values = [
                    f"'{branch['id']}'",
                    f"'{branch['name']}'",
                    f"'{branch['address']}'" if branch['address'] else "NULL",
                    f"'{branch['city']}'",
                    f"'{branch['state']}'",
                    f"'{branch['zip_code']}'",
                    f"'{branch['manager_id']}'" if branch['manager_id'] else "NULL",
                    f"'{branch['created_at']}'",
                    f"'{branch['updated_at']}'"
                ]
                f.write(f"INSERT INTO branches (id, name, address, city, state, zip_code, manager_id, created_at, updated_at) VALUES ({', '.join(values)});\n")
            
            f.write("\n-- Insert Employees\n")
            for employee in self.employees:
                values = [
                    f"'{employee['id']}'",
                    f"'{employee['branch_id']}'",
                    f"'{employee['name']}'",
                    f"'{employee['email']}'",
                    f"'{employee['phone']}'" if employee['phone'] else "NULL",
                    f"'{employee['position']}'",
                    f"'{employee['hire_date']}'",
                    f"{employee['salary']}",
                    f"'{employee['created_at']}'",
                    f"'{employee['updated_at']}'"
                ]
                f.write(f"INSERT INTO employees (id, branch_id, name, email, phone, position, hire_date, salary, created_at, updated_at) VALUES ({', '.join(values)});\n")
            
            f.write("\n-- Insert Customers\n")
            for customer in self.customers:
                values = [
                    f"'{customer['id']}'",
                    f"'{customer['email']}'",
                    f"'{customer['phone']}'" if customer['phone'] else "NULL",
                    f"'{customer['address']}'" if customer['address'] else "NULL",
                    f"'{customer['first_name']}'",
                    f"'{customer['last_name']}'",
                    f"'{customer['date_of_birth']}'",
                    f"'{customer['gender']}'",
                    f"'{customer['national_id']}'",
                    f"'{customer['created_at']}'",
                    f"'{customer['updated_at']}'",
                    f"'{customer['branch_id']}'"
                ]
                f.write(f"INSERT INTO customers (id, email, phone, address, first_name, last_name, date_of_birth, gender, national_id, created_at, updated_at, branch_id) VALUES ({', '.join(values)});\n")
            
            f.write("\n-- Insert Accounts\n")
            for account in self.accounts:
                values = [
                    f"'{account['id']}'",
                    f"'{account['customer_id']}'",
                    f"'{account['account_number']}'",
                    f"'{account['type']}'",
                    f"{account['balance']}",
                    f"'{account['opened_at']}'",
                    f"{account['interest_rate']}" if account['interest_rate'] else "NULL",
                    f"'{account['status']}'",
                    f"'{account['branch_id']}'",
                    f"'{account['created_at']}'",
                    f"'{account['updated_at']}'"
                ]
                f.write(f"INSERT INTO accounts (id, customer_id, account_number, type, balance, opened_at, interest_rate, status, branch_id, created_at, updated_at) VALUES ({', '.join(values)});\n")
            
            f.write("\n-- Insert Transactions\n")
            for transaction in self.transactions:
                values = [
                    f"'{transaction['id']}'",
                    f"'{transaction['account_id']}'",
                    f"'{transaction['transaction_date']}'",
                    f"{transaction['amount']}",
                    f"'{transaction['type']}'",
                    f"'{transaction['description']}'",
                    f"'{transaction['status']}'",
                    f"'{transaction['created_at']}'",
                    f"'{transaction['updated_at']}'",
                    f"'{transaction['employee_id']}'" if transaction['employee_id'] else "NULL"
                ]
                f.write(f"INSERT INTO transactions (id, account_id, transaction_date, amount, type, description, status, created_at, updated_at, employee_id) VALUES ({', '.join(values)});\n")
        
        print(f"SQL INSERT statements generated successfully!")
    
    def export_to_excel(self, filename="banking_test_data_now.xlsx"):
        """Export all data to Excel with separate sheets"""
        print(f"Exporting data to Excel file: {filename}...")
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Create DataFrames and export to different sheets
            branches_df = pd.DataFrame(self.branches)
            employees_df = pd.DataFrame(self.employees)
            customers_df = pd.DataFrame(self.customers)
            accounts_df = pd.DataFrame(self.accounts)
            transactions_df = pd.DataFrame(self.transactions)
            
            # Export each table to a separate sheet
            branches_df.to_excel(writer, sheet_name='Branches', index=False)
            employees_df.to_excel(writer, sheet_name='Employees', index=False)
            customers_df.to_excel(writer, sheet_name='Customers', index=False)
            accounts_df.to_excel(writer, sheet_name='Accounts', index=False)
            transactions_df.to_excel(writer, sheet_name='Transactions', index=False)
            
            # Create summary sheet
            summary_data = {
                'Table': ['Branches', 'Employees', 'Customers', 'Accounts', 'Transactions', 'Total'],
                'Record Count': [len(self.branches), len(self.employees), len(self.customers), 
                               len(self.accounts), len(self.transactions), 
                               len(self.branches) + len(self.employees) + len(self.customers) + len(self.accounts) + len(self.transactions)]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format the Excel file
        wb = openpyxl.load_workbook(filename)
        
        # Format headers
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        wb.save(filename)
        print(f"Excel file exported successfully!")

# Main execution
if __name__ == "__main__":
    # Generate test data
    generator = BankingDataGenerator()
    generator.generate_all_data()
    
    # Generate SQL inserts
    generator.generate_sql_inserts()
    
    # Export to Excel
    generator.export_to_excel()
    
    print("\n" + "="*50)
    print("TEST DATA GENERATION COMPLETED!")
    print("="*50)
    print("Files generated:")
    print("1. banking_test_data.sql - SQL INSERT statements")
    print("2. banking_test_data.xlsx - Excel file with separate sheets")
    print("\nEdge cases included:")
    print("- Very high/low balance accounts")
    print("- Old and new customer registrations")
    print("- Various transaction amounts and types")
    print("- Different account statuses")
    print("- Missing optional fields (phone, address)")
    print("- Complex relational data integrity")
